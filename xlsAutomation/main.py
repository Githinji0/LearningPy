import openpyxl as xl  # pip install openpyxl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

# cell = sheet["a1"]

sheet.cell(1, 1)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)

    # print(cell.value)

    new_Price = cell.value * 0.9
    # print(new_Price)
    newPriceCell = sheet.cell(row, 4)
    newPriceCell.value = new_Price
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=6)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
wb.save("transactions4.xlsx")
